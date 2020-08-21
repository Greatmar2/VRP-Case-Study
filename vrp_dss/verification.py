from time import perf_counter
from typing import List, Dict, Tuple

from openpyxl import load_workbook

from evaluate import evaluate_solution
from main import Runner
from model import run_settings
from settings import Data


def get_data_from_sheet(row: int) -> str:
    """Gets the output text file data from a certain row on the solve times summary sheet."""
    workbook = load_workbook(filename="Solve Times Summary.xlsx", read_only=True)
    run_data_sheet = workbook["Run Data"]
    return run_data_sheet[f"G{row}"].value


def extract_data_from_output(math_output: str) -> Tuple[Data, Dict[int, List[List[Tuple[int, int]]]]]:
    """Takes the mathematical output text and converts it to input data for the metaheuristic."""
    output_lines = math_output.split("\n")
    # Start after the "Input:" line
    current_line = 1

    # Loop through the customers and compile lists of information about them.
    locations: List[str] = ["Depot"]
    demands: List[int] = [0]
    window_starts: List[int] = [0]
    window_ends: List[int] = [24]
    average_unload_time: List[float] = [0.091667]
    coords: List[Tuple[float, float]] = [(0, 0)]
    while output_lines[current_line] and output_lines[current_line].count("Customer ") == 1:
        words = output_lines[current_line].split()
        # The fixed pattern of the output means the same words will always be in the same place
        # print([f"{num}: {word}" for num, word in enumerate(words)])
        # Example line: ['0: Customer', '1: 1', '2: has', '3: 5', '4: pallets', '5: demand', '6: and', '7: window',
        # '8: 0-24', '9: at', '10: (-5.625091957,', '11: 77.494351875)', '12: and', '13: average', '14: unload',
        # '15: time', '16: 0.120062083']
        locations.append(words[1])
        demands.append(int(words[3]))
        window = words[8].split("-")
        window_starts.append(int(window[0]))
        window_ends.append(int(window[1]))
        average_unload_time.append(float(words[16]))
        coords.append((float(words[10][1:-1]), float(words[11][:-1])))

        current_line += 1

    locations.append("DepotReturn")
    demands.append(0)
    window_starts.append(0)
    window_ends.append(32)
    average_unload_time.append(0)
    coords.append((0, 0))

    # Generate distances and times from Euclidean distances between the coordinates.
    distances = []
    times = []
    for index, from_loc in enumerate(coords):
        distances.append([])
        times.append([])
        for to_loc in coords:
            from_x, from_y = from_loc
            to_x, to_y = to_loc
            distance = ((from_x - to_x) ** 2 + (from_y - to_y) ** 2) ** 0.5
            distances[index].append(distance)
            times[index].append(distance / 80)

    # Loop through the vehicles to find information about the vehicle types.
    vehicle_name_types: Dict[str, int] = {}
    vehicle_types: List[str] = []
    distance_cost: List[float] = []
    time_cost: List[float] = []
    pallet_capacity: List[int] = []
    available_vehicles: List[int] = []
    while output_lines[current_line] and output_lines[current_line].count("Vehicle ") == 1:
        # print([f"{num}: {word}" for num, word in enumerate(words)])
        # ['0: Vehicle', '1: SP1', '2: is', '3: a', '4: 11', '5: metre', '6: with', '7: capacity', '8: 30,',
        # '9: distance', '10: cost', '11: 0.796243095,', '12: and', '13: time', '14: cost', '15: 10.888817567']
        # Find the vehicle type name
        line = output_lines[current_line]
        after_name = line.index(" with capacity")
        words: List[str] = line[:after_name].split()
        vehicle_type = " ".join(words[4:])
        vehicle_name = words[1]

        # Check if this vehicle type was already found
        if vehicle_types.count(vehicle_type) == 0:
            # Add the type to the list of types
            vehicle_types.append(vehicle_type)
            available_vehicles.append(1)
            # Get the properties of the vehicle
            words = line[after_name + 1:].split()
            pallet_capacity.append(int(words[2][:-1]))
            distance_cost.append(float(words[5][:-1]))
            time_cost.append(float(words[9]))
        else:
            # Increment the number of available vehicles of this type
            index = vehicle_types.index(vehicle_type)
            available_vehicles[index] += 1

        # Add the index of this vehicle to the dict that will later be used to identify the types of vehicles
        vehicle_name_types[vehicle_name] = vehicle_types.index(vehicle_type)

        current_line += 1

    run_data = Data(locations=locations, vehicle_types=vehicle_types, distance_cost=distance_cost, time_cost=time_cost,
                    pallet_capacity=pallet_capacity, available_vehicles=available_vehicles, demand=demands,
                    window_start=window_starts, window_end=window_ends, average_unload_time=average_unload_time,
                    distances=distances, times=times)

    # Skip blank and output header line
    current_line += 2

    # Iterate across solution lines use dictionary structure to compile a list of which vehicles travel where and to 
    # deliver how much, before compiling this into routes
    all_moves: Dict[str, Dict[str, Dict[str, str]]] = {}
    while output_lines[current_line] and output_lines[current_line].count("Vehicle ") == 1:
        words = output_lines[current_line].split()
        # print([f"{num}: {word}" for num, word in enumerate(words)])
        # ['0: Vehicle', '1: SP1', '2: travels', '3: from', '4: Depot', '5: to', '6: 7', '7: to', '8: deliver', '9: 5', 
        # '10: pallets.', '11: Expected', '12: unload', '13: start', '14: time', '15: is', '16: 5.084413751']
        vehicle = words[1]
        if not all_moves.get(vehicle):
            all_moves[vehicle] = {}
        from_loc = words[4]
        all_moves[vehicle][from_loc] = {"to": words[6], "load": words[9]}

        current_line += 1

    # Prepare the mathematical solution dict
    math_solution: Dict[int, List[List[Tuple[int, int]]]] = {}
    for type_index, _ in enumerate(vehicle_types):
        # Create the tour list
        math_solution[type_index] = []

    # Figure out what type each vehicle is, then compile the moves into routes.
    for vehicle, vehicle_moves in all_moves.items():
        # Find the type index for this vehicle
        type_index = vehicle_name_types[vehicle]
        route = []
        # Add the moves to the route
        move = vehicle_moves["Depot"]
        while move["to"] != "DepotReturn":
            route.append((int(move["to"]), int(move["load"])))
            move = vehicle_moves[move["to"]]

        # Add the route to the tour
        math_solution[type_index].append(route)

    return run_data, math_solution


def write_data_to_sheet(row: int, math_routes: str, math_objective: float, meta_routes: str, meta_time: float,
                        meta_objective: float):
    """Writes metaheuristic output data to the solve times summary sheet."""
    file_name = "Solve Times Summary.xlsx"
    workbook = load_workbook(filename=file_name)
    run_data_sheet = workbook["Run Data"]
    run_data_sheet[f"I{row}"].value = math_routes
    run_data_sheet[f"J{row}"].value = math_objective
    run_data_sheet[f"K{row}"].value = meta_routes
    run_data_sheet[f"L{row}"].value = meta_time
    run_data_sheet[f"M{row}"].value = meta_objective

    workbook.save(file_name)


if __name__ == "__main__":
    """Verify the metaheuristic against all mathematical instances."""
    start_row = 38  # The row to start on
    end_row = start_row + 1  # The row to stop before

    for row in range(start_row, end_row):
        text_data = get_data_from_sheet(row)
        input_data, exact_solution = extract_data_from_output(text_data)
        # print(input_data)
        # print(exact_solution)
        # print(input_data.repr_all())
        run_settings.set_run_data(input_data)
        # data_globals.update_globals()
        # model.run_data = input_data
        # model.run_config = Config()

        start_time = perf_counter()
        runner = Runner(5000, 1800, use_multiprocessing=False)
        best_solution = runner.run()
        end_time = perf_counter()

        print("Pretty output:\n" + best_solution.pretty_route_output())
        print(f"Run time: {end_time - start_time}")

        eval_results = evaluate_solution(exact_solution)

        write_data_to_sheet(row=row, math_routes=str(exact_solution), math_objective=eval_results["penalised_cost"],
                            meta_routes=best_solution.pretty_route_output(), meta_time=end_time - start_time,
                            meta_objective=best_solution.get_penalised_cost(1))
