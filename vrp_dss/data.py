"""This script provides functionality to import data and convert it to usable forms."""
import json
from math import inf, floor
from time import perf_counter
from typing import List, Optional, Tuple, Union, Dict

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from bing_key import api_key
from evaluate import evaluate_solution
from main import Runner
from model import run_settings
from settings import Data


class VehicleType:
    """Stores information about a vehicle type."""

    def __init__(self, data_index: int, name: str, distance_cost: float, time_cost: float, capacity: int,
                 vehicles_available: int):
        self.data_index = data_index
        self.name = name
        self.distance_cost = distance_cost
        self.time_cost = time_cost
        self.capacity = capacity
        self.vehicles_available = vehicles_available
        self.vehicles_used = 0

    def __eq__(self, other):
        if isinstance(other, VehicleType):
            return other.data_index == self.data_index or other.name == self.name
        if isinstance(other, Vehicle):
            return other.vehicle_type == self
        return False


class Vehicle:
    """Stores information about a vehicle."""

    def __init__(self, name: str, vehicle_type: VehicleType):
        self.name = name
        self.vehicle_type = vehicle_type

    def __eq__(self, other):
        if isinstance(other, Vehicle):
            return other.name == self.name
        return False


class PhysicalLocation:
    """Stores information about a real location serviced by SPAR. Each location can have a handful of customers,
    as SPAR, SUPERSPAR, KWIKSPAR, TOPS, and PHARMACY at the same location are treated as separate customers by SPAR.

    This script will treat stores at the same location as the same customer."""

    def __init__(self, data_index: int, name: str, latitude: float, longitude: float, offload: Union[str, float],
                 stores: str):
        """Initialise the object and convert the offload time and stores strings into appropriate values."""
        # super().__init__(data_index)
        self.data_index = data_index
        self.name = name
        self.latitude = latitude
        self.longitude = longitude
        if isinstance(offload, float):
            self.average_offload_time = offload
        else:
            self.average_offload_time = self.extract_average_offload(offload)
        self.store_ids = self.extract_store_ids(stores)
        self.demand = 0

    def __index__(self):
        return self.data_index

    def __repr__(self):
        return f"Location {self.name}, with store IDs {self.store_ids}."

    def __eq__(self, other):
        if isinstance(other, PhysicalLocation):
            return other.name == self.name
        if isinstance(other, int):
            return self.store_ids.count(other) > 0
        if isinstance(other, str):
            try:
                return self.store_ids.count(int(other)) > 0
            except (ValueError, TypeError):
                return False
        return False

    @property
    def anonymous_name(self) -> str:
        return f"{self.data_index}-{self.name[0]}"

    @staticmethod
    def extract_average_offload(offload_str: str) -> float:
        """
        Extracts the average offload time from the "target offload time" string.

        :param offload_str: String containing the target offload time and the target offload time.
        :returns: Integer representing the average offload time in hours.
        """
        timestamp = offload_str[offload_str.index("(") + 1:offload_str.index(")")]
        time_units = timestamp.split(":")
        offload_time = (int(time_units[0]) * 60 + int(time_units[1])) / 3600

        # If there is no average offload time stored, then set it to 5.5 minutes
        if offload_time == 0:
            offload_time = 0.09167

        return offload_time

    @staticmethod
    def extract_store_ids(stores_str: str) -> List[int]:
        """Extracts a list of the store IDs from the "stores at location" string.

        :param stores_str: String containing a list of stores at the location.
        :returns: A list of the IDs of stores at that location."""
        stores = stores_str[4:].split(", ")
        store_codes: List[int] = []
        for store in stores:
            store_code = store.split("-")[0]
            if store_code:
                store_codes.append(int(store_code))

        return store_codes


class Stop:
    def __init__(self, location: PhysicalLocation, delivered: int):
        self.location = location
        self.delivered = delivered
        # if delivered < 0:
        #     raise ValueError("Negative amount delivered!")


class Route:
    """Stores information from the archive about a route that a vehicle travelled."""

    def __init__(self, code: int, vehicle: Vehicle):
        self.code = code
        self.vehicle = vehicle
        self.stops: List[Stop] = []

    def __eq__(self, other):
        if isinstance(other, Route):
            return other.code == self.code
        if isinstance(other, int):
            return other == self.code
        return False

    def to_list(self) -> Tuple[int, List[Tuple[int, int]]]:
        """Returns a tuple containing list representing the route that can be used"""
        route: List[Tuple[int, int]] = [(stop.location.data_index, stop.delivered) for stop in self.stops]
        return self.vehicle.vehicle_type.data_index, route


def read_locations(filename: str = "Model Data.xlsx", worksheet_name: str = "Locations") -> List[
    PhysicalLocation]:
    """Reads in the store locations and creates a list of PhysicalLocation objects."""
    workbook = load_workbook(filename=filename, read_only=True)
    locations_sheet = workbook[worksheet_name]
    locations: List[PhysicalLocation] = [PhysicalLocation(0, "Depot", -34.005993, 18.537626, 0.09167, "")]
    row = 2
    while locations_sheet[f"A{row}"].value:
        locations.append(PhysicalLocation(data_index=row - 1, name=locations_sheet[f"A{row}"].value,
                                          latitude=locations_sheet[f"B{row}"].value,
                                          longitude=locations_sheet[f"C{row}"].value,
                                          offload=locations_sheet[f"E{row}"].value,
                                          stores=locations_sheet[f"F{row}"].value))
        row += 1
    return locations


def find_location(store_code: int, locations: List[PhysicalLocation]) -> Optional[PhysicalLocation]:
    """Searches the list of locations for one that contains the given store ID."""
    # Hopefully List.index can utilise the custom __eq__ function for PhysicalLocation.
    try:
        return locations[locations.index(store_code)]
    except ValueError:
        return None

    # for location in locations:
    #     if location.store_ids.count(store_code) > 0:
    #         return location

    # return None
    # raise ValueError(f"Store code {store_code} not found!")


def read_vehicles(vehicles_file: str, vehicles_sheet: str = "trucks", types_file: str = "Model Data.xlsx",
                  types_sheet: str = "Vehicle Types") -> Tuple[
    List[VehicleType], List[Vehicle]]:
    """Reads in information about the vehicle types and vehicles."""
    types_workbook = load_workbook(filename=types_file, read_only=True)
    types_sheet = types_workbook[types_sheet]
    vehicle_types: List[VehicleType] = []
    row = 2
    while types_sheet[f"A{row}"].value:
        vehicle_types.append(VehicleType(row - 2, name=types_sheet[f"A{row}"].value,
                                         distance_cost=types_sheet[f"B{row}"].value,
                                         time_cost=types_sheet[f"C{row}"].value,
                                         capacity=types_sheet[f"D{row}"].value,
                                         vehicles_available=types_sheet[f"E{row}"].value))
        row += 1

    vehicles_workbook = load_workbook(filename=vehicles_file, read_only=True)
    vehicles_sheet = vehicles_workbook[vehicles_sheet]
    vehicles: List[Vehicle] = []
    row = 2
    capacities = [vehicle_type.capacity for vehicle_type in vehicle_types]
    while vehicles_sheet[f"A{row}"].value:
        try:
            capacity = int(vehicles_sheet[f"C{row}"].value)
            vehicles.append(Vehicle(name=vehicles_sheet[f"A{row}"].value,
                                    vehicle_type=vehicle_types[capacities.index(capacity)]))
        except (ValueError, TypeError):
            pass
        row += 1

    return vehicle_types, vehicles


def find_vehicle(horse: str, trailer: str, carried: int, vehicles: List[Vehicle],
                 vehicle_types: List[VehicleType]) -> Optional[Vehicle]:
    """Searches the list of vehicles for one that contains the given vehicle name."""
    try:
        # Check if there is a vehicle that matches the horse code
        return vehicles[vehicles.index(horse)]
    except ValueError:
        try:
            # Check if there is a vehicle that matches the trailer code
            return vehicles[vehicles.index(trailer)]
        except ValueError:
            try:
                # Otherwise, try find the vehicle type that has the most similar to the capacity to the load carried
                lowest_diff = inf
                closest_type = None
                for vehicle_type in vehicle_types:
                    diff = abs(vehicle_type.capacity - carried)
                    if diff < lowest_diff:
                        lowest_diff = diff
                        closest_type = vehicle_type

                return Vehicle(horse, closest_type)

            except ValueError:
                return None


def find_route(route_code: str, routes: Dict[int, List[Route]]) -> Optional[Route]:
    """Searches the list of routes for one that contains the given route code."""
    for vehicle_type_index, tour in routes.items():
        count = tour.count(route_code)
        if count > 0:
            return tour[tour.index(route_code)]
        else:
            return None


def read_archive(filename: str, locations: List[PhysicalLocation], vehicles: List[Vehicle],
                 vehicle_types: List[VehicleType]) -> Dict[int, List[Route]]:
    """Reads in the delivery archive and sums up the demand per location and creating route lists."""
    workbook = load_workbook(filename=filename, read_only=True)
    deliveries_sheet = workbook["Deliveries"]
    routes: Dict[int, List[Route]] = {}
    row = 2
    route: Optional[Route] = None
    while deliveries_sheet[f"A{row}"].value:
        # Get the route code
        code = deliveries_sheet[f"D{row}"].value
        # If the route code does not match, then the vehicle is probably finished
        if route is None or code != route.code:
            if route:
                # In the archived data, vehicles are sometimes loaded with more than 30 pallets.
                # The schedulers make this work by consolidating pallets, reducing the actual number of pallets used.
                # This isn't handled by the algorithm, so - for the sake of fairness - any routes that have more pallets
                # than their vehicle's capacity will have their demand reduced.
                total_delivered = sum([stop.delivered for stop in route.stops])
                stop_ind = 0
                while total_delivered > route.vehicle.vehicle_type.capacity:
                    # Cycle through the stops, removing one demand from each at a time until total demand is acceptable
                    if route.stops[stop_ind].delivered > 1:
                        # Don't decrement the amount delivered to stops that have no more than 1 pallet delivered.
                        route.stops[stop_ind].delivered -= 1
                        route.stops[stop_ind].location.demand -= 1
                        total_delivered -= 1
                    stop_ind = (stop_ind + 1) % len(route.stops)
                if not routes.get(route.vehicle.vehicle_type.data_index):
                    routes[route.vehicle.vehicle_type.data_index] = []
                routes[route.vehicle.vehicle_type.data_index].append(route)
            # Check if the current route code matches anything already found
            # (just in case one vehicle's lines are spread out)
            # route = find_route(code, routes)
            # If not, create a new route
            # if not route:
            route = Route(code=code, vehicle=find_vehicle(horse=deliveries_sheet[f"J{row}"].value,
                                                          trailer=deliveries_sheet[f"K{row}"].value,
                                                          carried=deliveries_sheet[f"P{row}"].value,
                                                          vehicles=vehicles, vehicle_types=vehicle_types))
            route.vehicle.vehicle_type.vehicles_used += 1

        # Find the location
        location = find_location(deliveries_sheet[f"B{row}"].value, locations)

        if location:
            # Read in the demands, defaulting to 0 if they are blank or have "C/S"
            try:
                dry_demand = int(deliveries_sheet[f"M{row}"].value)
            except (ValueError, TypeError):
                dry_demand = 0
            try:
                perish_demand = int(deliveries_sheet[f"N{row}"].value)
            except (ValueError, TypeError):
                perish_demand = 0
            try:
                pick_by_line_demand = int(deliveries_sheet[f"O{row}"].value)
            except (ValueError, TypeError):
                pick_by_line_demand = 0

            # if dry_demand < 0 or perish_demand < 0 or pick_by_line_demand < 0:
            #     raise ValueError("Negative demand.")
            demand = dry_demand + perish_demand + pick_by_line_demand

            # Add the demand to the location's overall demand
            location.demand += demand
            # And set the amount delivered on this stop in the route
            # route.stops.append(Stop(location, demand))
            # Check whether the previous stop was at the same location. If so, merge them into one stop.
            if len(route.stops) > 0 and route.stops[0].location == location:
                route.stops[0].delivered += demand
            else:
                # Must be inserted at the start of the route, because the routes are in reverse order in the archives
                route.stops.insert(0, Stop(location, demand))

        row += 1

    return routes


def convert_routes_to_lists(routes: Dict[int, List[Route]]) -> Dict[int, List[List[Tuple[int, int]]]]:
    """Converts a dict of routes to lists of tuples."""
    lists: Dict[int, List[List[Tuple[int, int]]]] = {}
    for vehicle_type_index, tour in routes.items():
        lists[vehicle_type_index] = [[(stop.location.data_index, stop.delivered) for stop in route.stops] for route in
                                     tour]

    return lists


# def convert_lists_to_routes(lists: Dict[int, List[List[Tuple[int, int]]]]) -> Dict[int, List[Route]]:
#     """Converts a dict of lists of tuples to routes."""
#     routes: Dict[int, List[Route]] = {}
#     for vehicle_type_index, tuple_tour in lists.items():
#         tour = []
#         for tuple_route in tuple_tour:
#             route = Route()


def save_archive_routes(routes: Dict[int, List[Route]], filename: str = "Model Data.xlsx"):
    """Save the archive routes to the workbook in JSON format."""
    workbook = load_workbook(filename)
    workbook["Archive Routes"]["A1"].value = json.dumps(convert_routes_to_lists(routes))
    workbook.save(filename)


def load_archive_routes(filename: str = "Model Data.xlsx") -> Dict[int, List[List[List[int]]]]:
    """Load the archive routes from the workbook in JSON format."""
    workbook = load_workbook(filename, read_only=True)
    # Must convert the string keys to integers
    str_dict: Dict[str, List[List[List[int]]]] = json.loads(workbook["Archive Routes"]["A1"].value)
    int_dict: Dict[int, List[List[List[int]]]] = {}
    for vehicle_type, tour in str_dict.items():
        # Also remove stops that don't make any deliveries:
        for route in tour:
            index = 0
            while index < len(route):
                if route[index][1] == 0:
                    route.pop(index)

                index += 1

        int_dict[int(vehicle_type)] = tour
    return int_dict


def save_input_data(locations: List[PhysicalLocation], vehicle_types: List[VehicleType],
                    filename: str = "Model Data.xlsx", anonymised: bool = False):
    """Saves the information for all locations and vehicle types for the day imported."""
    workbook = load_workbook(filename)

    # Add the locations
    locations_sheet: Worksheet = workbook["Locations"]
    for index, location in enumerate(locations):
        if anonymised:
            locations_sheet[f"A{index + 2}"].value = location.anonymous_name
        else:
            locations_sheet[f"A{index + 2}"].value = location.name
            locations_sheet[f"B{index + 2}"].value = location.latitude
            locations_sheet[f"C{index + 2}"].value = location.longitude
        locations_sheet[f"D{index + 2}"].value = location.demand
        locations_sheet[f"E{index + 2}"].value = 0
        locations_sheet[f"F{index + 2}"].value = 24
        locations_sheet[f"G{index + 2}"].value = location.average_offload_time

    # Add the location distances and times
    # distances_sheet: Worksheet = workbook["Distances"]
    # times_sheet: Worksheet = workbook["Times"]
    # for row, distances_row in enumerate(distances):
    #     # Put location names at the start of each row
    #     if anonymised:
    #         distances_sheet.cell(row=row + 2, column=1, value=locations[row].anonymous_name)
    #         times_sheet.cell(row=row + 2, column=1, value=locations[row].anonymous_name)
    #     else:
    #         distances_sheet.cell(row=row + 2, column=1, value=locations[row].name)
    #         times_sheet.cell(row=row + 2, column=1, value=locations[row].name)
    #     for col, distance in enumerate(distances_row):
    #         if row == 0:
    #             # Put location names at the top of each column
    #             if anonymised:
    #                 distances_sheet.cell(row=1, column=col + 2, value=locations[col].anonymous_name)
    #                 times_sheet.cell(row=1, column=col + 2, value=locations[col].anonymous_name)
    #             else:
    #                 distances_sheet.cell(row=1, column=col + 2, value=locations[col].name)
    #                 times_sheet.cell(row=1, column=col + 2, value=locations[col].name)
    #         if row == col:
    #             # If on the same row and col, set the travel to be equal to going to the depot and back
    #             dist = distances[row][0] + distances[0][row]
    #             time = times[row][0] + times[0][row]
    #         else:
    #             dist = distances[row][col]
    #             time = times[row][col]
    #         # Apply the value to the sheet
    #         distances_sheet.cell(row=row + 2, column=col + 2, value=dist)
    #         times_sheet.cell(row=row + 2, column=col + 2, value=time)

    # Update the used vehicles counts
    vehicle_types_sheet: Worksheet = workbook["Vehicle Types"]
    for vehicle_type in vehicle_types:
        vehicle_types_sheet[f"F{vehicle_type.data_index + 2}"].value = vehicle_type.vehicles_used

    workbook.save(filename)


def convert_archive(archive_filename: str, data_filename: str = "Model Data.xlsx", anonymised: bool = False):
    """Reads in the demand, location, vehicle, and route data from the archive and then saves it."""
    locations = read_locations(filename="SPAR Locations and Schedule.xlsx", worksheet_name="Store Locations")
    vehicle_types, vehicles = read_vehicles(types_file=data_filename, types_sheet="Vehicle Types",
                                            vehicles_file="Spar Fleet.xlsx", vehicles_sheet="trucks")
    routes = read_archive(archive_filename, locations, vehicles, vehicle_types)
    save_input_data(locations, vehicle_types, anonymised=anonymised)
    save_archive_routes(routes)
    # if anonymised:
    #     location_names = [location.anonymous_name for location in locations]
    # else:
    #     location_names = [location.name for location in locations]


def import_data(filename: str = "Model Data.xlsx") -> Data:
    """Reads in the demand, location, and vehicle data to create input data for the algorithm.
    Also generates the routes from the archived data."""
    data_workbook = load_workbook(filename=filename, read_only=True, data_only=True)

    # Read in the location information
    location_sheet = data_workbook["Locations"]
    locations: List[str] = []
    demand: List[int] = []
    window_start: List[float] = []
    window_end: List[float] = []
    average_unload_time: List[float] = []
    row = 2
    while location_sheet[f"A{row}"].value:
        locations.append(location_sheet[f"A{row}"].value)
        demand.append(int(location_sheet[f"D{row}"].value))
        window_start.append(location_sheet[f"E{row}"].value)
        window_end.append(location_sheet[f"F{row}"].value)
        average_unload_time.append(location_sheet[f"G{row}"].value)

        row += 1

    # Read in the vehicle data
    vehicle_sheet = data_workbook["Vehicle Types"]
    vehicle_types: List[str] = []
    distance_cost: List[float] = []
    time_cost: List[float] = []
    pallet_capacity: List[int] = []
    available_vehicles: List[int] = []
    row = 2
    while vehicle_sheet[f"A{row}"].value:
        vehicle_types.append(vehicle_sheet[f"A{row}"].value)
        distance_cost.append(vehicle_sheet[f"B{row}"].value)
        time_cost.append(vehicle_sheet[f"C{row}"].value)
        pallet_capacity.append(int(vehicle_sheet[f"D{row}"].value))
        available_vehicles.append(int(vehicle_sheet[f"G{row}"].value))

        row += 1

    distances, times = import_matrix_input_data()

    archive_data = Data(locations=locations, demand=demand, window_start=window_start, window_end=window_end,
                        average_unload_time=average_unload_time, distances=distances, times=times,
                        vehicle_types=vehicle_types, distance_cost=distance_cost, time_cost=time_cost,
                        pallet_capacity=pallet_capacity, available_vehicles=available_vehicles)

    return archive_data


def pull_travel_data_from_bing(locations: List[PhysicalLocation]) -> Tuple[List[List[float]], List[List[float]]]:
    """Will generate a travel time and distance matrix between all locations using Bing Maps."""
    # Bing has a limit of 2500 origin-destination pairings for a distance matrix
    max_pairs = 2500
    # The locations will need to be iterated across with as many rows as possible at a time.
    rows_per_call = floor(max_pairs / len(locations))
    if rows_per_call == 0:
        raise ValueError("Too many locations!")
    start_row = 0
    distances: List[List[float]] = []
    times: List[List[float]] = []

    print(rows_per_call)

    # Keep iterating until all locations have been used as an origin to all other locations
    while start_row < len(locations):
        # Use a POST request to get information from the Bing maps API.
        # Example request and responses are found at
        # https://docs.microsoft.com/en-us/bingmaps/rest-services/examples/distance-matrix-example

        end_row = min(start_row + rows_per_call, len(locations))

        # Prepare the data for the post
        post_body = {"origins": [],
                     "destinations": [],
                     "travelMode": "driving"}
        # Request information from the current set of origins to all destinations
        for origin in locations[start_row:end_row]:
            post_body["origins"].append({"latitude": origin.latitude, "longitude": origin.longitude})
        for destination in locations:
            post_body["destinations"].append({"latitude": destination.latitude, "longitude": destination.longitude})
        print(post_body)

        # Send the request
        # Key in .gitignored file, because this repository is public.
        response = requests.post(f"https://dev.virtualearth.net/REST/v1/Routes/DistanceMatrix?key={api_key}",
                                 data=json.dumps(post_body))
        # The response is in JSON format
        # response: dict = json.loads(request.json())
        response_json: dict = response.json()
        if response_json["statusCode"] == 200:
            # Only interested in the results from the response
            results: List[Dict[str, Union[int, float]]] = response_json["resourceSets"][0]["resources"][0]["results"]
            result_index = 0

            # The results are a list of dicts, which iterate first through origins then destinations
            for origin_index, origin in enumerate(locations[start_row:end_row]):
                origin_distances = []
                origin_times = []
                for destination_index, destination in enumerate(locations):
                    # The response will exclude elements where the origin and destination are the same location
                    # if origin == destination:
                    #     continue

                    # The dict contains the origin and destination
                    if results[result_index]["destinationIndex"] != destination_index or \
                            results[result_index]["destinationIndex"] != destination_index:
                        raise ValueError(
                            f"Distance Matrix result indices at {result_index} don't match expected indices "
                            f"{origin_index} and {destination_index}.\n{results}")
                    # Store the expected travel distance and duration
                    origin_distances.append(results[result_index]["travelDistance"])
                    # Bing gives the durations in minutes, my algorithm uses hours
                    origin_times.append(results[result_index]["travelDuration"] / 60)

                    result_index += 1

                distances.append(origin_distances)
                times.append(origin_times)
            start_row = end_row
        else:
            raise ValueError(f"Request failed!\nRequest: {response.request}\nResponse: {response_json}")

    return distances, times


def import_matrix_input_data(filename: str = "Model Data.xlsx") -> Tuple[List[List[float]], List[List[float]]]:
    """Reads the matrices from the model data sheet."""
    workbook = load_workbook(filename, read_only=True, data_only=True)
    # distance_sheet: Worksheet = workbook["Distances"]
    # time_sheet: Worksheet = workbook["Times"]
    # row = 2
    # times: List[List[float]] = []
    # distances: List[List[float]] = []
    # while distance_sheet.cell(row=row, column=1).value:
    #     column = 2
    #     distance_row = []
    #     time_row = []
    #     while distance_sheet.cell(row=1, column=column).value:
    #         distance_row.append(distance_sheet.cell(row=row, column=column).value)
    #         time_row.append(time_sheet.cell(row=row, column=column).value)
    #
    #         column += 1
    #
    #     distances.append(distance_row)
    #     times.append(time_row)
    #     print(f"Row {row}")
    #
    #     row += 1

    # With the workbook opened in read only, it the worksheets are iterable objects
    distance_sheet: Worksheet = workbook["Distances"]
    time_sheet: Worksheet = workbook["Times"]
    distances: List[List[float]] = [[distance.value for distance in distance_row] for distance_row in distance_sheet]
    times: List[List[float]] = [[time.value for time in time_row] for time_row in time_sheet]

    # Can't use slices to cut the worksheets, so must cut the lists after taking the values from the sheets
    distances = [[distance for distance in distance_row[1:]] for distance_row in distances[1:]]
    times = [[time for time in time_row[1:]] for time_row in times[1:]]

    return distances, times


def save_matrix_input_data(locations: List[PhysicalLocation], distances: List[List[float]], times: List[List[float]],
                           filename: str = "Model Data.xlsx", anonymise: bool = False):
    """Saves the matrix information to the model data sheet."""
    workbook = load_workbook(filename)

    # Add the location distances and times
    distances_sheet: Worksheet = workbook["Distances"]
    times_sheet: Worksheet = workbook["Times"]
    for row, distances_row in enumerate(distances):
        # Put location names at the start of each row
        if anonymise:
            distances_sheet.cell(row=row + 2, column=1, value=locations[row].anonymous_name)
            times_sheet.cell(row=row + 2, column=1, value=locations[row].anonymous_name)
        else:
            distances_sheet.cell(row=row + 2, column=1, value=locations[row].name)
            times_sheet.cell(row=row + 2, column=1, value=locations[row].name)
        for col, distance in enumerate(distances_row):
            if row == 0:
                # Put location names at the top of each column
                if anonymise:
                    distances_sheet.cell(row=1, column=col + 2, value=locations[col].anonymous_name)
                    times_sheet.cell(row=1, column=col + 2, value=locations[col].anonymous_name)
                else:
                    distances_sheet.cell(row=1, column=col + 2, value=locations[col].name)
                    times_sheet.cell(row=1, column=col + 2, value=locations[col].name)
            if row == col:
                # If on the same row and col, set the travel to be equal to going to the depot and back
                dist = distances[row][0] + distances[0][row]
                time = times[row][0] + times[0][row]
            else:
                dist = distances[row][col]
                time = times[row][col]
            # Apply the value to the sheet
            distances_sheet.cell(row=row + 2, column=col + 2, value=dist)
            times_sheet.cell(row=row + 2, column=col + 2, value=time)

    workbook.save(filename)


def update_matrices(use_model_data_sheet: bool = True):
    """Reads in the location data from the appropriate sheet, requests the travel matrix from Bing, then saves this to
    the sheet."""
    if use_model_data_sheet:
        locations = read_locations()
    else:
        locations = read_locations("SPAR Locations and Schedule.xlsx", "Store Locations")

    distances, times = pull_travel_data_from_bing(locations)
    save_matrix_input_data(locations, distances, times, anonymise=True)


def save_output(filename: str, row: int = None, archive_routes: str = None, archive_cost: float = None,
                archive_penalty: float = None, meta_routes: str = None, meta_routes_pretty: str = None,
                meta_time: float = None, meta_cost: float = None, meta_penalty: float = None):
    """Writes metaheuristic output data to the solve times summary sheet."""
    workbook = load_workbook(filename=filename)
    run_data_sheet = workbook["Case Study"]
    if archive_routes:
        run_data_sheet[f"B{row}"].value = archive_routes
    if archive_cost:
        run_data_sheet[f"C{row}"].value = archive_cost
    if archive_penalty:
        run_data_sheet[f"D{row}"].value = archive_penalty
    if meta_routes:
        run_data_sheet[f"E{row}"].value = meta_routes
    if meta_routes_pretty:
        run_data_sheet[f"F{row}"].value = meta_routes_pretty
    if meta_time:
        run_data_sheet[f"G{row}"].value = meta_time
    if meta_cost:
        run_data_sheet[f"H{row}"].value = meta_cost
    if meta_penalty:
        run_data_sheet[f"I{row}"].value = meta_penalty

    workbook.save(filename)


def run_algorithm(output_row: int, output_filename: str = "Solve Times Summary.xlsx",
                  data_filename: str = "Model Data.xlsx"):
    """Imports data and runs the algorithm."""
    # Load the archive's routes to be compared
    # archive_routes = load_archive_routes()

    # Load data for this run
    print("Importing Data")
    run_data = import_data(data_filename)
    run_settings.set_run_data(run_data)

    # Run the algorithm, while timing it
    print("Starting Run")
    start_time = perf_counter()
    runner = Runner(10000, 3600, use_multiprocessing=False)
    best_solution = runner.run()
    end_time = perf_counter()

    # Evaluate the archive's solution
    # eval_results = evaluate_solution(archive_routes)

    print(f"Run time: {end_time - start_time}")
    if best_solution:
        print(f"Pretty output:\n{best_solution.pretty_route_output()}")
        # Save the results
        save_output(output_filename, row=output_row, meta_routes=json.dumps(best_solution.routes_to_dict()),
                    meta_routes_pretty=best_solution.pretty_route_output(), meta_time=end_time - start_time,
                    meta_cost=best_solution.cost, meta_penalty=best_solution.penalty)
    else:
        print(f"No feasible solution.")
        save_output(output_filename, row=output_row, meta_routes="None", meta_routes_pretty="None",
                    meta_time=end_time - start_time, meta_cost=0, meta_penalty=0)


def evaluate_archive_routes(output_row: int, output_filename: str = "Solve Times Summary.xlsx",
                            data_filename: str = "Model Data.xlsx"):
    """Loads archive's routes and evaluates them."""
    # Load the archive's routes to be compared
    archive_routes = load_archive_routes()

    # Load data for this run
    run_data = import_data(data_filename)
    run_settings.set_run_data(run_data)

    # Evaluate the archive's solution
    eval_results = evaluate_solution(archive_routes)

    save_output(output_filename, row=output_row, archive_routes=json.dumps(archive_routes),
                archive_cost=eval_results["cost"],
                archive_penalty=eval_results["capacity_penalty"] + eval_results["duration_penalty"])


if __name__ == "__main__":
    """Runs functions without the DSS GUI."""
    # Import the data from the archive and other sheets, then save it in the Model Data sheet.
    convert_archive("30 Oct 2019 Demands.xlsx", anonymised=True)
    # Update the travel matrix
    # update_matrices(False)

    # Call the algorithm to solve the problem
    # run_algorithm(2)
    # Evaluate the original solution to the problem
    # evaluate_archive_routes(2)
