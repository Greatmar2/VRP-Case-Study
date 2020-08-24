"""This script provides functionality to import data and convert it to usable forms."""
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from settings import Data


class PhysicalLocation:
    """Stores information about a real location serviced by SPAR. Each location can have a handful of customers,
    as SPAR, SUPERSPAR, KWIKSPAR, TOPS, and PHARMACY at the same location are treated as separate customers by SPAR.

    This script will treat stores at the same location as the same customer."""

    def __init__(self, data_index: int, name: str, latitude: float, longitude: float, offload_str: str,
                 stores_str: str):
        """Initialise the object and convert the offload time and stores strings into appropriate values."""
        # super().__init__(data_index)
        self.data_index = data_index
        self.name = name
        self.latitude = latitude
        self.longitude = longitude
        self.average_offload_time = self.extract_average_offload(offload_str)
        self.store_ids = self.extract_store_ids(stores_str)
        self.demand = 0

    def __index__(self):
        return self.data_index

    def __repr__(self):
        return f"Location {self.name}, with store IDs {self.store_ids}."

    def __eq__(self, other):
        if isinstance(other, PhysicalLocation):
            return other.name == self.name
        if isinstance(other, str):
            return self.store_ids.count(other) > 0
        return False

    @staticmethod
    def extract_average_offload(offload_str: str) -> int:
        """
        Extracts the average offload time from the "target offload time" string.

        :param offload_str: String containing the target offload time and the target offload time.
        :returns: Integer representing the average offload time in seconds.
        """
        # Expected offload string format: 00:00  (16:58)   min/container
        timestamp = offload_str[offload_str.index("(") + 1:offload_str.index(")")]
        time_units = timestamp.split(":")
        offload_time = int(time_units[0]) * 60 + int(time_units[1])

        return offload_time

    @staticmethod
    def extract_store_ids(stores_str: str) -> List[str]:
        """Extracts a list of the store IDs from the "stores at location" string.

        :param stores_str: String containing a list of stores at the location.
        :returns: A list of the IDs of stores at that location."""
        # Expected store string format: (3) 35668-Aurora SPAR, 35805-Aurora TOPS, 35924-Aurora PHARMACY
        stores = stores_str[4:].split(", ")
        store_ids: List[str] = []
        for store in stores:
            store_ids.append(store.split("-")[0])

        return store_ids


class VehicleType:
    """Stores information about a vehicle type."""

    def __init__(self, data_index: int, name: str, distance_cost: float, time_cost: float):
        self.data_index = data_index
        self.name = name
        self.distance_cost = distance_cost
        self.time_cost = time_cost


class Vehicle:
    """Stores information about a vehicle."""

    def __init__(self, name: str, vehicle_type: VehicleType):
        self.name = name
        self.vehicle_type = vehicle_type


class Route:
    """Stores information from the archive about a route that a vehicle travelled."""

    def __init__(self, route_code: int, vehicle: Vehicle):
        self.route_code = route_code
        self.vehicle = vehicle
        self.stops: List[PhysicalLocation] = []
        self.loads: List[int] = []

    def __eq__(self, other):
        if isinstance(other, Route):
            return other.route_code == self.route_code
        if isinstance(other, int):
            return other == self.route_code
        return False

    def add_stop(self, stop: PhysicalLocation, load: int):
        self.stops.append(stop)
        self.loads.append(load)

    def get_route_list(self) -> Tuple[int, List[Tuple[int, int]]]:
        """Returns a tuple containing list representing the route that can be used"""
        route: List[Tuple[int, int]] = [(stop.data_index, load) for stop, load in zip(self.stops, self.loads)]
        return self.vehicle.vehicle_type.data_index, route


def read_locations() -> List[PhysicalLocation]:
    """Reads in the store locations and creates a list of PhysicalLocation objects."""
    workbook = load_workbook(filename="SPAR Locations and Schedule.xlsx")
    locations_sheet = workbook["Store Locations"]
    locations: List[PhysicalLocation] = []
    row = 2
    while locations_sheet[f"A{row}"].value:
        locations.append(PhysicalLocation(data_index=row - 2, name=locations_sheet[f"A{row}"].value,
                                          latitude=locations_sheet[f"B{row}"].value,
                                          longitude=locations_sheet[f"C{row}"].value,
                                          offload_str=locations_sheet[f"E{row}"].value,
                                          stores_str=locations_sheet[f"F{row}"].value))
    return locations


def find_location(store_id: str, locations: List[PhysicalLocation]) -> Optional[PhysicalLocation]:
    """Searches the list of locations for one that contains the given store ID."""
    # Hopefully List.index can utilise the custom __eq__ function for PhysicalLocation.
    # Otherwise I'll use a list comprehension.
    try:
        return locations[locations.index(store_id)]
    except ValueError:
        return None


def summarise_archive(filename: str, locations: List[PhysicalLocation], vehicles: List[Vehicle],
                      vehicle_types: List[VehicleType]) -> List[Route]:
    """Reads in the delivery archive and sums up the demand per location and creating route lists."""
    workbook = load_workbook(filename=filename)
    deliveries_sheet = workbook["Deliveries"]
    routes: List[Route] = []
    row = 2
    while deliveries_sheet[f"A{row}"].value:
        location = find_location(deliveries_sheet[f"A{row}"].value, locations)
        # Read in the demands, defaulting to 0 if they are blank or have "C/S"
        try:
            dry_demand = int(deliveries_sheet[f"M{row}"].value)
        except ValueError:
            dry_demand = 0
        try:
            perish_demand = deliveries_sheet[f"N{row}"].value
        except ValueError:
            perish_demand = 0
        try:
            pick_by_line_demand = deliveries_sheet[f"O{row}"].value
        except ValueError:
            pick_by_line_demand = 0


def import_data(demand_file: str) -> Data:
    """Reads in the demand, location, and vehicle data to create input data for the algorithm."""
