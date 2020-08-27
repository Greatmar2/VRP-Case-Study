from math import floor
from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from data import PhysicalLocation


def pull_travel_data_from_bing(locations: List[PhysicalLocation]) -> Tuple[List[List[float]], List[List[float]]]:
    """Will generate a travel time and distance matrix between all locations using Bing Maps."""
    # Bing has a limit of 2500 origin-destination pairings for a distance matrix
    max_pairs = 2500
    # The locations will need to be iterated across with as many rows as possible at a time.
    rows_per_call = floor(max_pairs / len(locations))
    start_row = 0
    post_body = {"origins": [],
                 "destinations": [],
                 "travelMost": "driving"}


def read_input_data() -> Tuple[List[List[float]], List[List[float]]]:
    """Reads the matrices from the model data sheet."""
    workbook = load_workbook("Model Data.xlsx", read_only=True)
    distance_sheet: Worksheet = workbook["Distances"]
    time_sheet: Worksheet = workbook["Times"]
    row = 2
    column = 2
    times: List[List[float]] = []
    distances: List[List[float]] = []
    while distance_sheet.cell(row=row, column=1).value:
        distances.append([])
        times.append([])
        while distance_sheet.cell(row=1, column=column).value:
            distances.append(distance_sheet.cell(row=row, column=column).value)
            times.append(time_sheet.cell(row=row, column=column).value)

            column += 1
        row += 1

    return distances, times


def save_input_data(locations: List[PhysicalLocation], distances: List[List[float]], times: List[List[float]],
                    anonymise: bool = False):
    """Saves the matrix information to the model data sheet."""
    workbook = load_workbook("Model Data.xlsx")

    # Add the location distances and times
    distances_sheet: Worksheet = workbook["Distances"]
    times_sheet: Worksheet = workbook["Times"]
    for row, distances_row in enumerate(distances):
        # Put location names at the start of each row
        if anonymise:
            distances_sheet.cell(row=row + 2, column=1, value=locations[row].anonymous_name)
            times_sheet.cell(row=row + 2, column=1, value=locations[row].anonymous_name)
        else:
            distances_sheet.cell(row=row + 2, column=1, value=locations[row].name)
            times_sheet.cell(row=row + 2, column=1, value=locations[row].name)
        for col, distance in enumerate(distances_row):
            if row == 0:
                # Put location names at the top of each column
                if anonymise:
                    distances_sheet.cell(row=1, column=col + 2, value=locations[col].anonymous_name)
                    times_sheet.cell(row=1, column=col + 2, value=locations[col].anonymous_name)
                else:
                    distances_sheet.cell(row=1, column=col + 2, value=locations[col].name)
                    times_sheet.cell(row=1, column=col + 2, value=locations[col].name)
            if row == col:
                # If on the same row and col, set the travel to be equal to going to the depot and back
                dist = distances[row][0] + distances[0][row]
                time = times[row][0] + times[0][row]
            else:
                dist = distances[row][col]
                time = times[row][col]
            # Apply the value to the sheet
            distances_sheet.cell(row=row + 2, column=col + 2, value=dist)
            times_sheet.cell(row=row + 2, column=col + 2, value=time)

    workbook.save("Model Data.xlsx")
